#!/usr/bin/env python3
"""Build the bounded mixed-format Stage 5 hero collection.

The generated directory is an integration fixture, not a source of benchmark
labels.  Text and binary inputs deliberately repeat a small vocabulary so
native extractors and measured Stage 4 signals have real cross-source evidence.
No adapter needs to expose raw rows or arrays for the collection to be useful.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import json
import os
import re
import sqlite3
import tempfile
import zipfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
HERO = HERE / "hero" / "borealis"
MANIFEST = HERE / "hero" / "manifest.json"
FIXED_ZIP_TIME = (2026, 8, 30, 0, 0, 0)
FROZEN_CORE_MODIFIED = b"2026-08-30T00:00:00Z"
CORE_MODIFIED_PATTERN = re.compile(
    rb"(<dcterms:modified[^>]*>)[^<]+(</dcterms:modified>)"
)
RAW_ONLY_XLSX_SENTINEL = "AUTOTLDR_RAW_ONLY_XLSX_20260830_5A1D2C"
RAW_ONLY_PARQUET_SENTINEL = "AUTOTLDR_RAW_ONLY_PARQUET_20260830_8E4B17"
RAW_ONLY_SQLITE_SENTINEL = "AUTOTLDR_RAW_ONLY_SQLITE_20260830_7F3C91"
RAW_ONLY_DUCKDB_SENTINEL = "AUTOTLDR_RAW_ONLY_DUCKDB_20260830_C2A845"
RAW_ONLY_HDF5_SENTINEL = "AUTOTLDR_RAW_ONLY_HDF5_20260830_D6E913"
RAW_ONLY_NETCDF_SENTINEL = "AUTOTLDR_RAW_ONLY_NETCDF_20260830_B4F720"


def _write_text(name: str, content: str) -> None:
    path = HERO / name
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8", newline="\n")


def _write_json(name: str, value: object) -> None:
    _write_text(name, json.dumps(value, indent=2, sort_keys=True) + "\n")


def _build_text_sources() -> None:
    _write_text(
        "overview.md",
        """# Borealis reservoir safety

Borealis monitors and controls the Station Alpha cooling reservoir. The
operating target `reservoir_temp_c` is 18.0 °C and the pressure ceiling
`pressure_kpa` is 240 kPa. `controller.py` reads `config.json`, consumes
`measurements.parquet`, and records safety events in `safety.sqlite`.

Scientific context is split between `experiments.h5` and `forecast.nc`.
`analytics.duckdb` stores bounded aggregate profiles, while `capacity.xlsx`
computes the reserve and `safety_margin_pct`. Operators follow
`operations.html` and review `pipeline.ipynb`.

The current calibration dependency is `calibration/current.csv`; that file is
referenced by the operating notes but is not included in this collection.
""",
    )
    _write_json(
        "config.json",
        {
            "station_id": "alpha",
            "reservoir_temp_c": {"target": 18.0, "unit": "degC"},
            "pressure_kpa": {"ceiling": 240.0, "unit": "kPa"},
            "safety_margin_pct": 12.0,
            "telemetry_source": "measurements.parquet",
            "event_store": "safety.sqlite",
        },
    )
    _write_text(
        "controller.py",
        '''"""Borealis reservoir safety controller."""

from pathlib import Path

CONFIG_PATH = Path("config.json")
TELEMETRY_PATH = Path("measurements.parquet")
EVENT_STORE = Path("safety.sqlite")


def safety_state(
    reservoir_temp_c: float,
    pressure_kpa: float,
    target_temp_c: float = 18.0,
    pressure_ceiling_kpa: float = 240.0,
) -> str:
    """Return alert when reservoir temperature or pressure exceeds policy."""
    if reservoir_temp_c > target_temp_c or pressure_kpa > pressure_ceiling_kpa:
        return "alert"
    return "nominal"
''',
    )
    _write_text(
        "operations.html",
        """<!doctype html>
<html lang="en"><head><title>Borealis operations</title></head><body>
<h1>Borealis operating procedure</h1>
<p>The Station Alpha reservoir cooling service keeps reservoir_temp_c near
18.0 degC and pressure_kpa below 240 kPa.</p>
<ol><li>Read config.json before starting controller.py.</li>
<li>Validate measurements.parquet and the calibration/current.csv dependency.</li>
<li>Write alerts to safety.sqlite and review capacity.xlsx reserve margin.</li></ol>
<p>experiments.h5 and forecast.nc provide scientific context; neither is an
operator command source.</p>
</body></html>
""",
    )
    _write_json(
        "pipeline.ipynb",
        {
            "cells": [
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "# Borealis validation\n",
                        "Compare measurements.parquet with forecast.nc and experiments.h5.",
                    ],
                },
                {
                    "cell_type": "code",
                    "execution_count": 1,
                    "metadata": {},
                    "outputs": [
                        {
                            "name": "stdout",
                            "output_type": "stream",
                            "text": ["station alpha validation complete\n"],
                        }
                    ],
                    "source": [
                        "RESERVOIR_TEMP_C = 'reservoir_temp_c'\n",
                        "PRESSURE_KPA = 'pressure_kpa'\n",
                        "print('station alpha validation complete')\n",
                    ],
                },
            ],
            "metadata": {
                "kernelspec": {
                    "display_name": "Python 3",
                    "language": "python",
                    "name": "python3",
                }
            },
            "nbformat": 4,
            "nbformat_minor": 5,
        },
    )
    _write_text(
        "station.csv",
        "metric,unit,source\nreservoir_temp_c,degC,measurements.parquet\n"
        "pressure_kpa,kPa,measurements.parquet\nsafety_margin_pct,percent,capacity.xlsx\n",
    )


def _canonicalize_xlsx(raw: Path, target: Path) -> None:
    members: list[tuple[str, bytes]] = []
    with zipfile.ZipFile(raw, "r") as source:
        for name in sorted(source.namelist()):
            payload = source.read(name)
            if name == "docProps/core.xml":
                payload, replacements = CORE_MODIFIED_PATTERN.subn(
                    rb"\g<1>" + FROZEN_CORE_MODIFIED + rb"\g<2>", payload
                )
                if replacements != 1:
                    raise RuntimeError("XLSX core modified timestamp was not unique")
            members.append((name, payload))
    temporary = target.with_suffix(target.suffix + ".tmp")
    with zipfile.ZipFile(
        temporary,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
        strict_timestamps=True,
    ) as output:
        for name, payload in members:
            info = zipfile.ZipInfo(name, date_time=FIXED_ZIP_TIME)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o100644 << 16
            output.writestr(
                info,
                payload,
                compress_type=zipfile.ZIP_DEFLATED,
                compresslevel=9,
            )
    os.replace(temporary, target)


def _build_xlsx() -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    workbook.properties.creator = "AutoTLDR Stage 5 benchmark"
    workbook.properties.created = dt.datetime(2026, 8, 30, tzinfo=dt.timezone.utc)
    workbook.properties.modified = dt.datetime(2026, 8, 30, tzinfo=dt.timezone.utc)
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs.append(("parameter", "value", "unit"))
    inputs.append(("reservoir_temp_c", 18.0, "degC"))
    inputs.append(("pressure_kpa", 240.0, "kPa"))
    inputs.append(("safety_margin_pct", 12.0, "percent"))
    capacity = workbook.create_sheet("Capacity")
    capacity.append(("metric", "formula"))
    capacity.append(("temperature_reserve_c", "=Inputs!B2-15"))
    capacity.append(("pressure_reserve_kpa", "=Inputs!B3-210"))
    capacity.append(("effective_margin_pct", "=Inputs!B4"))
    raw_canary = workbook.create_sheet("_raw_canary")
    raw_canary.sheet_state = "veryHidden"
    raw_canary["A1"] = RAW_ONLY_XLSX_SENTINEL
    with tempfile.TemporaryDirectory(prefix="autotldr-stage5-xlsx-") as raw_dir:
        raw = Path(raw_dir) / "capacity.raw.xlsx"
        workbook.save(raw)
        workbook.close()
        _canonicalize_xlsx(raw, HERO / "capacity.xlsx")


def _build_parquet() -> None:
    import pyarrow as pa
    import pyarrow.parquet as pq

    schema = pa.schema(
        [
            ("timestamp_utc", pa.timestamp("s", tz="UTC")),
            ("reservoir_temp_c", pa.float64()),
            ("pressure_kpa", pa.float64()),
            ("station_id", pa.string()),
            ("raw_note", pa.string()),
        ],
        metadata={
            b"dataset": b"Borealis Station Alpha telemetry",
            b"reservoir_temp_c.unit": b"degC",
            b"pressure_kpa.unit": b"kPa",
        },
    )
    table = pa.Table.from_arrays(
        [
            pa.array(
                [
                    dt.datetime(2026, 8, 30, 12, 0, tzinfo=dt.timezone.utc),
                    dt.datetime(2026, 8, 30, 12, 5, tzinfo=dt.timezone.utc),
                    dt.datetime(2026, 8, 30, 12, 10, tzinfo=dt.timezone.utc),
                ],
                type=schema.field("timestamp_utc").type,
            ),
            pa.array([17.8, 18.1, 18.0], type=pa.float64()),
            pa.array([218.0, 221.0, 219.5], type=pa.float64()),
            pa.array(["alpha", "alpha", "alpha"], type=pa.string()),
            pa.array(
                [RAW_ONLY_PARQUET_SENTINEL, None, None], type=pa.string()
            ),
        ],
        schema=schema,
    )
    pq.write_table(
        table,
        HERO / "measurements.parquet",
        compression="NONE",
        version="2.6",
        write_statistics=True,
    )


def _build_sqlite() -> None:
    target = HERO / "safety.sqlite"
    temporary = target.with_suffix(".sqlite.tmp")
    temporary.unlink(missing_ok=True)
    connection = sqlite3.connect(temporary)
    try:
        connection.executescript(
            """
            PRAGMA page_size = 4096;
            PRAGMA foreign_keys = ON;
            CREATE TABLE stations (
                station_id TEXT PRIMARY KEY,
                description TEXT NOT NULL
            );
            CREATE TABLE safety_events (
                event_id INTEGER PRIMARY KEY,
                station_id TEXT NOT NULL REFERENCES stations(station_id),
                reservoir_temp_c REAL NOT NULL,
                pressure_kpa REAL NOT NULL,
                state TEXT NOT NULL,
                observed_at TEXT NOT NULL
            );
            CREATE INDEX safety_events_station ON safety_events(station_id);
            """
        )
        connection.execute(
            "INSERT INTO stations VALUES (?, ?)",
            ("alpha", RAW_ONLY_SQLITE_SENTINEL),
        )
        connection.execute(
            "INSERT INTO safety_events VALUES (?, ?, ?, ?, ?, ?)",
            (1, "alpha", 18.1, 221.0, "nominal", "2026-08-30T12:05:00Z"),
        )
        connection.commit()
        connection.execute("VACUUM")
    finally:
        connection.close()
    os.replace(temporary, target)


def _build_duckdb() -> None:
    import duckdb

    target = HERO / "analytics.duckdb"
    temporary = target.with_suffix(".duckdb.tmp")
    temporary.unlink(missing_ok=True)
    connection = duckdb.connect(str(temporary))
    try:
        connection.execute(
            """
            CREATE TABLE telemetry_profile (
                station_id VARCHAR PRIMARY KEY,
                reservoir_temp_c_mean DOUBLE,
                pressure_kpa_mean DOUBLE,
                sample_count BIGINT,
                raw_note VARCHAR
            );
            """
        )
        connection.execute(
            "INSERT INTO telemetry_profile VALUES (?, ?, ?, ?, ?)",
            ["alpha", 17.9666667, 219.5, 3, RAW_ONLY_DUCKDB_SENTINEL],
        )
        connection.execute(
            """
            CREATE VIEW reservoir_headroom AS
            SELECT station_id, 18.0 - reservoir_temp_c_mean AS temp_headroom_c,
                   240.0 - pressure_kpa_mean AS pressure_headroom_kpa
            FROM telemetry_profile;
            """
        )
        connection.execute("CHECKPOINT")
    finally:
        connection.close()
    os.replace(temporary, target)


def _build_hdf5() -> None:
    import h5py

    target = HERO / "experiments.h5"
    temporary = target.with_suffix(".h5.tmp")
    temporary.unlink(missing_ok=True)
    with h5py.File(temporary, "w", track_order=True, libver="earliest") as output:
        output.attrs["title"] = "Borealis reservoir cooling experiments"
        output.attrs["station_id"] = "alpha"
        run = output.create_group("experiments/run_001", track_order=True)
        run.attrs["method"] = "closed-loop cooling validation"
        temperature = run.create_dataset(
            "reservoir_temp_c",
            data=[17.9, 18.0, 18.1],
            track_times=False,
        )
        temperature.attrs["units"] = "degC"
        pressure = run.create_dataset(
            "pressure_kpa",
            data=[218.0, 219.5, 221.0],
            track_times=False,
        )
        pressure.attrs["units"] = "kPa"
        run.create_dataset(
            "raw_note",
            data=RAW_ONLY_HDF5_SENTINEL.encode("ascii"),
            dtype=f"S{len(RAW_ONLY_HDF5_SENTINEL)}",
            track_times=False,
        )
    os.replace(temporary, target)


def _build_netcdf() -> None:
    import netCDF4
    import numpy as np

    target = HERO / "forecast.nc"
    temporary = target.with_suffix(".nc.tmp")
    temporary.unlink(missing_ok=True)
    output = netCDF4.Dataset(temporary, "w", format="NETCDF4_CLASSIC")
    try:
        output.title = "Borealis Station Alpha reservoir forecast"
        output.createDimension("time", 3)
        time = output.createVariable("time", "i4", ("time",))
        time.units = "minutes since 2026-08-30 12:00:00 UTC"
        time[:] = [0, 5, 10]
        temperature = output.createVariable("reservoir_temp_c", "f8", ("time",))
        temperature.units = "degC"
        temperature.long_name = "forecast reservoir temperature"
        temperature[:] = [17.9, 18.0, 18.2]
        pressure = output.createVariable("pressure_kpa", "f8", ("time",))
        pressure.units = "kPa"
        pressure.long_name = "forecast reservoir pressure"
        pressure[:] = [219.0, 220.0, 222.0]
        output.createDimension("raw_note_char", len(RAW_ONLY_NETCDF_SENTINEL))
        raw_note = output.createVariable("raw_note", "S1", ("raw_note_char",))
        raw_note[:] = np.frombuffer(
            RAW_ONLY_NETCDF_SENTINEL.encode("ascii"), dtype="S1"
        )
    finally:
        output.close()
    os.replace(temporary, target)


def read_raw_canaries(root: Path | None = None) -> dict[str, str]:
    """Read each Tier-3 canary through its native library and exact address."""

    import duckdb
    import h5py
    import netCDF4
    import openpyxl
    import pyarrow.parquet as pq

    source_root = root or HERO
    workbook = openpyxl.load_workbook(
        source_root / "capacity.xlsx", data_only=True, read_only=False
    )
    try:
        sheet = workbook["_raw_canary"]
        if sheet.sheet_state != "veryHidden":
            raise RuntimeError("hero XLSX raw-only canary sheet is not veryHidden")
        xlsx_value = sheet["A1"].value
    finally:
        workbook.close()

    parquet = pq.read_table(source_root / "measurements.parquet", columns=["raw_note"])
    parquet_value = parquet.column("raw_note")[0].as_py()

    sqlite_path = source_root / "safety.sqlite"
    sqlite_uri = sqlite_path.resolve().as_uri() + "?mode=ro&immutable=1"
    with sqlite3.connect(sqlite_uri, uri=True) as connection:
        sqlite_row = connection.execute(
            "SELECT description FROM stations WHERE station_id = ?",
            ("alpha",),
        ).fetchone()

    duckdb_connection = duckdb.connect(
        database=str(source_root / "analytics.duckdb"), read_only=True
    )
    try:
        duckdb_row = duckdb_connection.execute(
            "SELECT raw_note FROM telemetry_profile WHERE station_id = ?",
            ["alpha"],
        ).fetchone()
    finally:
        duckdb_connection.close()

    with h5py.File(source_root / "experiments.h5", "r") as hdf5:
        hdf5_value = hdf5["/experiments/run_001/raw_note"][()]

    netcdf = netCDF4.Dataset(source_root / "forecast.nc", mode="r")
    try:
        netcdf_value = netcdf.variables["raw_note"][:].tobytes().decode("ascii")
    finally:
        netcdf.close()

    if sqlite_row is None or duckdb_row is None:
        raise RuntimeError("hero database raw-only canary row is absent")
    return {
        "capacity.xlsx": str(xlsx_value),
        "measurements.parquet": str(parquet_value),
        "safety.sqlite": str(sqlite_row[0]),
        "analytics.duckdb": str(duckdb_row[0]),
        "experiments.h5": bytes(hdf5_value).decode("ascii"),
        "forecast.nc": netcdf_value,
    }


def _build_archive() -> None:
    target = HERO / "appendix.zip"
    temporary = target.with_suffix(".zip.tmp")
    members = {
        "appendix/calibration.md": (
            "# Calibration note\n\nThe controller expects calibration/current.csv "
            "for Station Alpha sensor calibration. The current file is absent.\n"
        ).encode("utf-8"),
        "appendix/ownership.json": json.dumps(
            {"calibration_owner": "operations", "station_id": "alpha"},
            indent=2,
            sort_keys=True,
        ).encode("utf-8")
        + b"\n",
    }
    with zipfile.ZipFile(
        temporary,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
        strict_timestamps=True,
    ) as output:
        for name in sorted(members):
            info = zipfile.ZipInfo(name, date_time=FIXED_ZIP_TIME)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o100644 << 16
            output.writestr(info, members[name], compresslevel=9)
    os.replace(temporary, target)


def _manifest() -> dict[str, object]:
    files = []
    for path in sorted(item for item in HERO.rglob("*") if item.is_file()):
        payload = path.read_bytes()
        files.append(
            {
                "path": path.relative_to(HERO).as_posix(),
                "bytes": len(payload),
                "sha256": hashlib.sha256(payload).hexdigest(),
            }
        )
    return {
        "schema": 1,
        "collection": "borealis-stage5-hero-v1",
        "generated_with": {
            "python": "3.12+",
            "binary_libraries": {
                "duckdb": "runtime-recorded",
                "h5py": "runtime-recorded",
                "netCDF4": "runtime-recorded",
                "openpyxl": "runtime-recorded",
                "pyarrow": "runtime-recorded",
            },
        },
        "files": files,
    }


def _write_manifest() -> None:
    MANIFEST.parent.mkdir(parents=True, exist_ok=True)
    MANIFEST.write_text(
        json.dumps(_manifest(), indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
        newline="\n",
    )


def _build_locked_tier3() -> None:
    _build_xlsx()
    _build_parquet()
    _build_sqlite()
    _build_duckdb()
    _build_hdf5()
    _build_netcdf()


def build() -> None:
    HERO.mkdir(parents=True, exist_ok=True)
    _build_text_sources()
    _build_locked_tier3()
    _build_archive()
    _write_manifest()


def rebuild_tier3() -> None:
    """Rebuild only the locked Tier-3 fixtures and collection manifest."""

    _build_locked_tier3()
    _write_manifest()


def validate() -> None:
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    expected = manifest.get("files")
    if not isinstance(expected, list):
        raise RuntimeError("hero manifest has no files list")
    observed = _manifest()["files"]
    if expected != observed:
        raise RuntimeError("hero fixture bytes differ from manifest; rebuild explicitly")
    required = {
        ".md",
        ".py",
        ".json",
        ".csv",
        ".html",
        ".ipynb",
        ".zip",
        ".xlsx",
        ".parquet",
        ".sqlite",
        ".duckdb",
        ".h5",
        ".nc",
    }
    present = {path.suffix for path in HERO.iterdir() if path.is_file()}
    missing = sorted(required - present)
    if missing:
        raise RuntimeError(f"hero fixture is missing required suffixes: {missing}")
    expected_canaries = {
        "capacity.xlsx": RAW_ONLY_XLSX_SENTINEL,
        "measurements.parquet": RAW_ONLY_PARQUET_SENTINEL,
        "safety.sqlite": RAW_ONLY_SQLITE_SENTINEL,
        "analytics.duckdb": RAW_ONLY_DUCKDB_SENTINEL,
        "experiments.h5": RAW_ONLY_HDF5_SENTINEL,
        "forecast.nc": RAW_ONLY_NETCDF_SENTINEL,
    }
    if read_raw_canaries() != expected_canaries:
        raise RuntimeError("one or more hero Tier-3 raw-only canaries are absent or changed")


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("command", choices=("build", "rebuild-tier3", "validate"))
    args = parser.parse_args()
    if args.command == "build":
        build()
    elif args.command == "rebuild-tier3":
        rebuild_tier3()
    validate()
    print(json.dumps({"status": "valid", "root": str(HERO), "files": len(_manifest()["files"])}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
