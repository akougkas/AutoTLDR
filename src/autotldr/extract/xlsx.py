"""Spreadsheets, read as a formula dependency graph.

A spreadsheet's meaning is not its values. It is what depends on what. Which
cells are free inputs, which are derived, what the derivation is, and where
somebody typed a number directly into a formula instead of referencing an
assumption. Routing an XLSX through a markdown converter throws all of that away
and leaves a grid of numbers, which is why the router never does.

Nothing in this module calls a model.
"""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from ..unit import Extraction, Modality, Origin, Relation, RelationKind, Role, Unit

# Cell and range references, with optional sheet qualifier and $ anchors.
_REF = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
    r"(?P<start>\$?[A-Z]{1,3}\$?\d{1,7})"
    r"(?::(?P<end>\$?[A-Z]{1,3}\$?\d{1,7}))?"
)
# A bare number inside a formula, ignoring those glued to a cell ref or an
# exponent. These are the buried overrides worth surfacing.
_LITERAL = re.compile(r"(?<![A-Za-z0-9_$.:])(\d+\.?\d*)(?![A-Za-z0-9_(])")
_FUNC = re.compile(r"([A-Z][A-Z0-9.]{1,30})\s*\(")

# Literals that carry no analytical weight and would only add noise.
_BENIGN_LITERALS = frozenset({"0", "1", "2", "-1", "100", "0.0", "1.0"})


def extract(path: Path) -> Extraction:
    source = str(path)
    result = Extraction(source=source, kind="xlsx")

    formulas = openpyxl.load_workbook(path, data_only=False, read_only=False)
    try:
        values = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:  # pragma: no cover - some writers omit the value cache
        values = None

    try:
        _extract_workbook(result, formulas, values, source)
    finally:
        formulas.close()
        if values is not None:
            values.close()

    return result


def _extract_workbook(result: Extraction, wb, values_wb, source: str) -> None:
    unit_by_cell: dict[str, Unit] = {}
    deps: dict[str, set[str]] = {}
    referenced: set[str] = set()
    formula_cells: dict[str, str] = {}
    labels: dict[str, str] = {}

    # Pass 1: read every cell, record formulas and their dependencies.
    for ws in wb.worksheets:
        grid, texts = _read_sheet(ws)
        result.units.append(_sheet_unit(source, ws, grid))

        for addr, raw in grid.items():
            if isinstance(raw, str) and raw.startswith("="):
                key = f"{ws.title}!{addr}"
                formula_cells[key] = raw
                targets = _dependencies(raw, ws.title)
                deps[key] = targets
                referenced |= targets

        labels.update(_infer_labels(ws.title, grid, texts))

    # Pass 2: emit a unit per formula cell, and per constant that a formula uses.
    for key, formula in formula_cells.items():
        sheet, addr = key.split("!", 1)
        value = _cached_value(values_wb, sheet, addr)
        label = labels.get(key, "")
        literals = _suspect_literals(formula)

        unit = Unit(
            source=source,
            modality=Modality.RECORD,
            content=_describe(label, addr, formula, value),
            origin=Origin(source, key),
            role=Role.RESULT,
            structure=(sheet,),
            salience=0.75 if key not in referenced else 0.6,
            meta={
                "cell": key,
                "formula": formula,
                "value": value,
                "label": label or None,
                "functions": sorted(_FUNC.findall(formula.upper())),
                "hardcoded": literals or None,
                "terminal": key not in referenced,
            },
        )
        result.units.append(unit)
        unit_by_cell[key] = unit

        if literals:
            result.gaps.append(
                f"{key} hardcodes {', '.join(literals)} inside its formula "
                f"instead of referencing a named input"
            )

    for key in sorted(referenced - set(formula_cells)):
        sheet, addr = key.split("!", 1)
        value = _cached_value(values_wb, sheet, addr)
        if value is None:
            continue
        label = labels.get(key, "")
        unit = Unit(
            source=source,
            modality=Modality.RECORD,
            content=_describe(label, addr, None, value),
            origin=Origin(source, key),
            role=Role.ASSUMPTION,
            structure=(sheet,),
            salience=0.85,
            meta={
                "cell": key,
                "value": value,
                "label": label or None,
                "input": True,
                "documented": bool(label),
            },
        )
        result.units.append(unit)
        unit_by_cell[key] = unit

    # Pass 3: turn the dependency map into relations.
    for key, targets in deps.items():
        src_unit = unit_by_cell.get(key)
        if src_unit is None:
            continue
        for target in sorted(targets):
            if dst_unit := unit_by_cell.get(target):
                result.relations.append(
                    Relation(
                        src=src_unit.id,
                        dst=dst_unit.id,
                        kind=RelationKind.DERIVES_FROM,
                        evidence=f"{key} formula references {target}",
                    )
                )

    _report_findings(result, deps, referenced, formula_cells, labels, unit_by_cell)

    result.meta.update(
        {
            "sheets": [ws.title for ws in wb.worksheets],
            "formula_cells": len(formula_cells),
            "input_cells": sum(1 for u in result.units if u.meta.get("input")),
        }
    )


# --------------------------------------------------------------------------- #
# Sheet reading
# --------------------------------------------------------------------------- #


def _read_sheet(ws) -> tuple[dict[str, object], dict[str, str]]:
    """Return every non-empty cell, plus the subset holding text."""
    grid: dict[str, object] = {}
    texts: dict[str, str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            addr = f"{get_column_letter(cell.column)}{cell.row}"
            grid[addr] = cell.value
            if isinstance(cell.value, str) and not cell.value.startswith("="):
                texts[addr] = cell.value.strip()
    return grid, texts


def _sheet_unit(source: str, ws, grid: dict[str, object]) -> Unit:
    formulas = sum(1 for v in grid.values() if isinstance(v, str) and v.startswith("="))
    summary = (
        f"Sheet {ws.title!r}: {ws.max_row} rows x {ws.max_column} columns, "
        f"{len(grid)} populated cells, {formulas} formulas."
    )
    return Unit(
        source=source,
        modality=Modality.SCHEMA,
        content=summary,
        origin=Origin(source, f"{ws.title}!A1:{get_column_letter(max(ws.max_column, 1))}{ws.max_row}"),
        role=Role.DEFINITION,
        structure=(ws.title,),
        salience=0.9,
        meta={
            "sheet": ws.title,
            "rows": ws.max_row,
            "columns": ws.max_column,
            "populated": len(grid),
            "formulas": formulas,
        },
    )


# --------------------------------------------------------------------------- #
# Formula analysis
# --------------------------------------------------------------------------- #


def _dependencies(formula: str, current_sheet: str) -> set[str]:
    """Every cell a formula reads, ranges expanded to their corners.

    Expanding a range fully would produce thousands of edges for a
    ``SUM(A1:A5000)`` and tell you nothing extra, so ranges contribute their
    endpoints and record the span in the relation's evidence instead.
    """
    found: set[str] = set()
    body = formula[1:] if formula.startswith("=") else formula

    for m in _REF.finditer(body):
        sheet = (m.group("sheet") or current_sheet).strip("'")
        for group in ("start", "end"):
            if ref := m.group(group):
                found.add(f"{sheet}!{ref.replace('$', '')}")
    return found


def _suspect_literals(formula: str) -> list[str]:
    """Numbers typed directly into a formula that probably should be inputs."""
    out: list[str] = []
    for raw in _LITERAL.findall(formula):
        if raw in _BENIGN_LITERALS:
            continue
        out.append(raw)
    return sorted(set(out))


def _infer_labels(sheet: str, grid: dict[str, object], texts: dict[str, str]) -> dict[str, str]:
    """Best-effort human label for each cell.

    Spreadsheets label by convention rather than by structure: the text
    immediately to the left of a value, or failing that the column header above
    it. Both conventions are common and neither is guaranteed, so an unlabeled
    cell is reported as undocumented rather than given a made-up name.
    """
    labels: dict[str, str] = {}
    by_col: dict[str, list[int]] = defaultdict(list)
    for addr in texts:
        col, row = _split(addr)
        by_col[col].append(row)

    for addr in grid:
        col, row = _split(addr)
        if left := _nearest_left(texts, col, row):
            labels[f"{sheet}!{addr}"] = left
            continue
        if header := _column_header(texts, by_col, col, row):
            labels[f"{sheet}!{addr}"] = header
    return labels


def _split(addr: str) -> tuple[str, int]:
    i = 0
    while i < len(addr) and addr[i].isalpha():
        i += 1
    return addr[:i], int(addr[i:] or 0)


def _col_index(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    return n


def _nearest_left(texts: dict[str, str], col: str, row: int) -> str:
    """Scan up to three columns left for a text label in the same row."""
    idx = _col_index(col)
    for back in range(1, 4):
        if idx - back < 1:
            break
        candidate = f"{get_column_letter(idx - back)}{row}"
        if label := texts.get(candidate):
            return label
    return ""


def _column_header(texts: dict[str, str], by_col: dict[str, list[int]], col: str, row: int) -> str:
    """The nearest text cell above, in the same column."""
    rows = [r for r in by_col.get(col, ()) if r < row]
    if not rows:
        return ""
    return texts.get(f"{col}{max(rows)}", "")


def _cached_value(values_wb, sheet: str, addr: str):
    if values_wb is None or sheet not in values_wb.sheetnames:
        return None
    try:
        value = values_wb[sheet][addr].value
    except (KeyError, ValueError, IndexError):  # pragma: no cover
        return None
    return value if not isinstance(value, str) or not value.startswith("=") else None


def _describe(label: str, addr: str, formula: str | None, value) -> str:
    name = label or addr
    if formula:
        # The formula carries its own leading '=', so do not add a second one.
        shown = f"{name} {formula}" if formula.startswith("=") else f"{name} = {formula}"
        return f"{shown} -> {value}" if value is not None else shown
    return f"{name} = {value}"


# --------------------------------------------------------------------------- #
# Findings
# --------------------------------------------------------------------------- #


def _report_findings(
    result: Extraction,
    deps: dict[str, set[str]],
    referenced: set[str],
    formula_cells: dict[str, str],
    labels: dict[str, str],
    unit_by_cell: dict[str, Unit],
) -> None:
    undocumented = [
        u.meta["cell"]
        for u in result.units
        if u.meta.get("input") and not u.meta.get("documented")
    ]
    if undocumented:
        shown = ", ".join(sorted(undocumented)[:5])
        more = f" and {len(undocumented) - 5} more" if len(undocumented) > 5 else ""
        result.gaps.append(f"input cells with no adjacent label: {shown}{more}")

    if cycles := _cycles(deps):
        for cycle in cycles[:5]:
            result.gaps.append("circular reference: " + " -> ".join(cycle))

    outputs = [key for key in formula_cells if key not in referenced]
    if not outputs and formula_cells:
        result.gaps.append("every formula feeds another; the workbook has no terminal output")


def _cycles(deps: dict[str, set[str]]) -> list[list[str]]:
    """Find dependency cycles by iterative depth-first search.

    Iterative rather than recursive because a deep chain in a large workbook
    would otherwise blow the stack on exactly the sheets most worth analyzing.
    """
    found: list[list[str]] = []
    state: dict[str, int] = {}  # 0 = visiting, 1 = done

    for root in deps:
        if state.get(root) == 1:
            continue
        stack: list[tuple[str, list[str]]] = [(root, [root])]
        while stack:
            node, path = stack.pop()
            if state.get(node) == 1:
                continue
            state[node] = 0
            for nxt in sorted(deps.get(node, ())):
                if nxt in path:
                    found.append(path[path.index(nxt):] + [nxt])
                    continue
                if state.get(nxt) != 1 and nxt in deps:
                    stack.append((nxt, path + [nxt]))
            state[node] = 1
    return found
